from flask import Flask, request, jsonify, Blueprint, render_template, abort
from flask_login import current_user, login_required
import traceback
from utilities.extensions import db
from models.contracts.contracts import CPContractStage, CPContractsImport,CPContract
from models.clients.clients import CPClientStage, CPClient
from functools import wraps


DEVMODE=True
contracts = Blueprint('contracts', __name__, template_folder='templates')

from openpyxl import load_workbook
from flask import jsonify

def process_row_nexus(row, current_contract, contracts, contracts_cpf):
    contract_id = "N" + str(row[0])
    name, cpf, unidades = row[1], row[4], row[2]
    if name and not isinstance(name, int) and len(contract_id) > 1 and contract_id != "NNone" and name != "NOME":
        # New contract or additional clients under current contract
        if current_contract is None or current_contract['contrato'] != contract_id:
            # Start a new contract record
            contracts_cpf = []
            current_contract = {
                'contrato': contract_id,
                'clientes': [],
                'unidades': ''
            }

        # Create a unique identifier for each client
        if cpf and cpf not in contracts_cpf:
            current_contract['clientes'].append({'nome': name, 'cpf': cpf})
            contracts_cpf.append(cpf)
    
    if isinstance(name, str) and "UNIDADES:" in name and current_contract:
        current_contract['unidades'] = unidades
        contracts.append(current_contract)

    return current_contract, contracts, contracts_cpf

def process_row_mega(row, current_contract, contracts, contracts_cpf):
    if isinstance(row[4], int) or 'Participante' in str(row[0]):
        contract_id = "M" + str(row[4])
        if contract_id != current_contract and str(row[4]) != 'None':
            contracts.append(current_contract)
            contracts_cpf = []
            current_contract = {
                'contrato': contract_id,
                'clientes': []
            }

        name = row[2] if str(row[1]).replace('0', '').isdigit() else row[1]
        client = {'nome': name}
        current_contract['clientes'].append(client)
        
    if 'CPF' in str(row[0]) or 'CNPJ' in str(row[0]):
        cpf = row[1]
        if cpf and cpf not in contracts_cpf:
            current_contract['clientes'][-1]['cpf']=cpf
            contracts_cpf.append(cpf)
            
    return current_contract, contracts, contracts_cpf

def process_row_uau(row, current_contract, contracts, contracts_cpf):
    if isinstance(row[16], int) and row[14] and 'Código' in row[14]:
        contract_id = "U" + str(row[16])
        if contract_id != current_contract and str(row[16]) != 'None':
            contracts.append(current_contract)
            contracts_cpf = []
            current_contract = {
                'contrato': contract_id,
                'unidades' : '',
                'clientes': []
            }
            
        if row[0] and 'Cliente' in row[0]:
            name = row[3] 
            client = {'nome': name}
            current_contract['clientes'].append(client)
        
    if row[0] and 'CPF' in str(row[0]) or 'CNPJ' in str(row[0]):
        cpf = row[3]
        if cpf and cpf not in contracts_cpf:
            current_contract['clientes'][0]['cpf']=cpf
            contracts_cpf.append(cpf)
    
    if  row[3] and 'VENDA' in row[3]:
        if not current_contract.get('unidades'):
            current_contract['unidades'] = f"{str(row[6])} - {str(row[7])}"
        else:
            current_contract['unidades'] =  current_contract['unidades'] + f" - {str(row[6])} - {str(row[7])}"
            
    return current_contract, contracts, contracts_cpf

def read_excel_nexus(sheet):
    contracts, current_contract, contracts_cpf = [], None, []
    
    for row in sheet.iter_rows(min_row=7, values_only=True):
        current_contract, contracts, contracts_cpf = process_row_nexus(row, current_contract, contracts, contracts_cpf)
    return contracts

def read_excel_mega(sheet):
    contracts, current_contract, contracts_cpf = [], None, []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        current_contract, contracts, contracts_cpf = process_row_mega(row, current_contract, contracts, contracts_cpf)
    return contracts

def read_excel_uau(sheet):
    contracts, current_contract, contracts_cpf = [], None, []
    for row in sheet.iter_rows(min_row=6, values_only=True):
        current_contract, contracts, contracts_cpf = process_row_uau(row, current_contract, contracts, contracts_cpf)
        
    return contracts

def read_excel_file(file):
    #function to read excel file
    workbook = load_workbook(filename=file)
    sheet = workbook.active

    #for nexus workbook it will be start at row 6
    headers_nexus_uau = [cell.value for cell in sheet[6]]

    #for mega workbook it will be start at row 1
    header_mega = [cell.value for cell in sheet[1]]
    #check if its NEXUS.
    if 'CÓDIGO' in headers_nexus_uau[0]:
        return read_excel_nexus(sheet)
    #check if its MEGA.
    if 'Unidade' in header_mega:
        return read_excel_mega(sheet)
    #check if its UAU.
    if 'Cliente' in headers_nexus_uau[0]:
        return read_excel_uau(sheet)
    
    
    return jsonify({"error": "Arquivo inválido!"}), 401

def is_file_already_imported(file):
    import_control = CPContractsImport.query.filter_by(file_name=file.filename).first()
    return import_control and not DEVMODE

def handle_transaction(db_operation, *args):
    try:
        db_operation(*args)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        error_message = f"Error during database operation. Contact the administrator: {str(e)}"
        return jsonify({"error": error_message}), 500

def create_import_control_row(file):
    """
    Creates an import control row for the given file.

    Args:
    file (str): The file name.

    Returns:
    CPContractsImport: The import control row.
    Error: An error message if the file has already been imported.
    """

    if is_file_already_imported(file):
        return jsonify({"error": "File already imported!"}), 402

    import_control = CPContractsImport(file_name=file.filename)
    err = handle_transaction(db.session.add, import_control)
    # If an error occurred and was handled in handle_transaction, early return here
    if err:
        return err
    
    return import_control

def insert_contract_data(list_contracts, import_control):
    """
    Creates clients and contracts in the Stage Table.

    Args:
    list_contracts (list of dicts): The data of clients to process.
    import_control (CPContractsImport): The import control row.

    Returns:
    bool: True if processing was successful, False otherwise.
    Error: An error message if CPF is invalid.
    """
    if list_contracts[0]==None:
        list_contracts.pop(0)
 
    try:
        for contract in list_contracts:
            create_records_for_each_contract(contract, import_control)
        
        import_control.total_clients_imported = len(list_contracts)
        db.session.commit()

        return True
    except Exception as e:
        handle_exception(e)

def create_records_for_each_contract(contract, import_control):
    contract_stage = CPContractStage(
        contract_code=contract['contrato'],
        unidades=contract.get('unidades'),
    )

    db.session.add(contract_stage)
    create_records_for_each_client(contract['clientes'], import_control, contract['contrato'])
    db.session.commit()

def create_records_for_each_client(client_list, import_control, contract_code):
    for client in client_list:
        client_stage = CPClientStage(
            name=client['nome'],
            cpf_cnpj=client['cpf'],
            import_id=import_control.id,
            contract_code=contract_code
        )
        db.session.add(client_stage)

def handle_exception(err):
    print(err)
    db.session.rollback()
    error_message = "Error while inserting Clients. Contact the Administrator. " + str(err) + "\n" + traceback.format_exc()
    return jsonify({"Error": error_message}), 400

@contracts.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file is sent."}), 400

    file = request.files['file']
    import_control = create_import_control_row(file)
    if not is_successful_import_control(import_control):
        return import_control

    list_contracts = read_excel_file(file)

    if not has_clients_in_contracts(list_contracts):
        return list_contracts

    create_contracts = insert_contract_data(list_contracts, import_control)
    if not is_successful_create_contracts(create_contracts):
        return list_contracts

    mark_import_control_ready(import_control)
    return jsonify({"success": True, "message": "Imported to Stage finish"}), 200


def is_successful_import_control(import_control):
    return not isinstance(import_control, tuple)


def has_clients_in_contracts(list_contracts):
    return list_contracts[1].get('clientes')


def is_successful_create_contracts(create_contracts):
    return not isinstance(create_contracts, tuple)


def mark_import_control_ready(import_control):
    import_control.status="Pronto para Importar"
    db.session.commit()

@contracts.route('/home')
def home():
    return render_template('contracts/home.html')

@contracts.route('/importList')
def import_contract():
    data = db.session.query(CPContractsImport).all()
    count = len(data)
    return render_template('contracts/home_import_list.html', count=count, data=data)

@contracts.route('/importProcessStep0')
def import_process_step0():
    data = db.session.query(CPContractsImport).all()
    count = len(data)
    return render_template('contracts/import_process_view0.html', count=count, data=data)
